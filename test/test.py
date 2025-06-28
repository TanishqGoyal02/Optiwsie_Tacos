import openpyxl
import os

import pandas as pd

def read_excel_cell_a6_from_sheet(filename, sheet_name="By Variants"):
    """
    Read the content of cell A6 from a specific sheet in an Excel file and print it.
    
    Args:
        filename (str): Name of the Excel file to read
        sheet_name (str): Name of the sheet to read from (default: "By Variants")
    """
    try:
        # Check if file exists
        if not os.path.exists(filename):
            print(f"Error: File '{filename}' not found in the current directory.")
            return
        
        # Load the workbook
        workbook = openpyxl.load_workbook(filename)
        
        # Check if the specified sheet exists
        if sheet_name not in workbook.sheetnames:
            print(f"Error: Sheet '{sheet_name}' not found in the workbook.")
            print(f"Available sheets: {workbook.sheetnames}")
            workbook.close()
            return
        
        # Get the specified sheet
        sheet = workbook[sheet_name]
        
        # Method 1: Using pandas with openpyxl engine (what you're currently using)
        print("=== Method 1: pandas with openpyxl engine ===")
        df1 = pd.read_excel(filename, sheet_name=sheet_name, engine="openpyxl")
        print(df1.head())
        
        # Method 2: Using pandas with xlrd engine (for older Excel files)
        print("\n=== Method 2: pandas with xlrd engine ===")
        try:
            df2 = pd.read_excel(filename, sheet_name=sheet_name, engine="xlrd")
            print(df2.head())
        except Exception as e:
            print(f"xlrd engine failed: {e}")
        
        # Method 3: Using pandas with odf engine (for OpenDocument files)
        print("\n=== Method 3: pandas with odf engine ===")
        try:
            df3 = pd.read_excel(filename, sheet_name=sheet_name, engine="odf")
            print(df3.head())
        except Exception as e:
            print(f"odf engine failed: {e}")
        
        # Method 4: Using pandas without specifying engine (auto-detect)
        print("\n=== Method 4: pandas auto-detect engine ===")
        df4 = pd.read_excel(filename, sheet_name=sheet_name)
        print(df4.head())
        
        # Method 6: Read all sheets at once
        print("\n=== Method 6: Read all sheets ===")
        all_sheets = pd.read_excel(filename, sheet_name=None, engine="openpyxl")
        for sheet_name_key, df_sheet in all_sheets.items():
            print(f"Sheet: {sheet_name_key}, Shape: {df_sheet.shape}")
        
        # Process sheet with hyperlink handling
        print("\n=== Processing sheet with hyperlink extraction ===")
        rows = []
        for row in sheet.iter_rows():
            row_data = []
            for cell in row:
                val = cell.value
                if isinstance(val, str) and val.startswith("=HYPERLINK("):
                    # Extract the display text: second argument
                    inside = val[len("=HYPERLINK("):-1]  # remove =HYPERLINK( and ending )
                    parts = inside.split(",", 1)
                    if len(parts) == 2:
                        display = parts[1].strip().strip('"')
                        row_data.append(display)
                    else:
                        row_data.append(val)
                else:
                    row_data.append(val)
            rows.append(row_data)
        
        # Create DataFrame from processed data
        df_processed = pd.DataFrame(rows)
        print("DataFrame with hyperlink processing:")
        print(df_processed.head())
        
        # Close the workbook
        workbook.close()
        
    except Exception as e:
        print(f"Error reading Excel file: {e}")

if __name__ == "__main__":
    # Read the Excel file from specific sheet and cell
    read_excel_cell_a6_from_sheet("/Users/tanishqgoyal/Documents/talk-to-tacos/test7.xlsx", "By Variants")
