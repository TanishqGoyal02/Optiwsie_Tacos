import openpyxl

def create_sample_excel():
    """
    Create a sample Excel file named 'text.xlsx' with a sheet called 'By Variants' and content in cell A6
    """
    # Create a new workbook
    workbook = openpyxl.Workbook()
    
    # Get the active sheet and rename it to "By Variants"
    sheet = workbook.active
    sheet.title = "By Variants"
    
    # Add some content to cell A6
    sheet['A6'] = "This is the content from cell A6 in the By Variants sheet!"
    
    # Add some additional content to other cells for demonstration
    sheet['A1'] = "Header 1"
    sheet['A2'] = "Header 2"
    sheet['A3'] = "Header 3"
    sheet['A4'] = "Header 4"
    sheet['A5'] = "Header 5"
    sheet['B6'] = "This is cell B6"
    sheet['C6'] = "This is cell C6"
    
    # Create another sheet for demonstration
    sheet2 = workbook.create_sheet("Sheet2")
    sheet2['A1'] = "This is Sheet2 content"
    
    # Save the workbook
    workbook.save('text.xlsx')
    
    print("Sample Excel file 'text.xlsx' created successfully!")
    print("Sheet 'By Variants' created with content in cell A6: 'This is the content from cell A6 in the By Variants sheet!'")
    print("Available sheets:", workbook.sheetnames)

if __name__ == "__main__":
    create_sample_excel() 