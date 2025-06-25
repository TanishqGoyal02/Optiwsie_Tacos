from openpyxl import load_workbook
import pandas as pd
from sqlalchemy import create_engine, text
import os
import numpy as np
import re

def detect_header_row(path, sheet_name, expected_keywords, min_matches=2):
    """
    Detects the header row in a given sheet by matching expected keywords.
    Returns 0-based row index if found, else raises ValueError.
    """
    wb = load_workbook(path, read_only=True)
    ws = wb[sheet_name]
    
    for i, row in enumerate(ws.iter_rows(values_only=True), start=0):
        row_values = [str(cell).strip().lower() if cell is not None else "" for cell in row]
        match_count = sum(any(expected.lower() in cell for cell in row_values) for expected in expected_keywords)
        if match_count >= min_matches:
            print(f"✅ Detected header row in '{sheet_name}' at index: {i}")
            return i

    raise ValueError(f"❌ Header row not found in sheet '{sheet_name}' — expected keywords not matched.")

def create_db_from_excel(excel_path, db_path):
    """
    Reads an Excel file and creates a SQLite database from its sheets.
    Each sheet is converted into a separate table with proper header handling.
    """
    if os.path.exists(db_path):
        os.remove(db_path)

    xls = pd.ExcelFile(excel_path)
    engine = create_engine(f"sqlite:///{db_path}")

    for sheet_name in xls.sheet_names:
        sheet_name_lower = sheet_name.lower()

        try:
            # Summary
            if "summary" in sheet_name_lower:
                df = pd.read_excel(xls, sheet_name=sheet_name)
                print(f"Processing summary table '{sheet_name}' with standard headers")

            # By Item IDs
            elif "by item ids" in sheet_name_lower:
                header_index = detect_header_row(excel_path, sheet_name, ["Item ID", "Product Name", "SKU"])
                df = pd.read_excel(excel_path, sheet_name=sheet_name, header=header_index)

                # Process first two columns for HYPERLINKs
                df.iloc[:, 0:2] = df.iloc[:, 0:2].astype(str)
                hyperlink_pattern = r'=HYPERLINK\s*\(\s*"[^"]*"\s*,\s*"([^"]*)"\s*\)'
                for col in df.columns[:2]:
                    df[col] = df[col].apply(lambda x: 
                        re.search(hyperlink_pattern, str(x)).group(1) 
                        if isinstance(x, str) and re.search(hyperlink_pattern, str(x)) 
                        else x)
                print(f"Processed table '{sheet_name}' with hyperlink extraction")

            # By Keywords
            elif "by keywords" in sheet_name_lower:
                header_index = detect_header_row(excel_path, sheet_name, ["Keyword"], min_matches=1)
                df = pd.read_excel(excel_path, sheet_name=sheet_name, header=header_index)
                print(f"Processing keyword table '{sheet_name}'")

            # Trend - Period
            elif "trend - period" in sheet_name_lower:
                header_index = detect_header_row(excel_path, sheet_name, ["Item Id", "Product Name", "SKU"])
                df_all = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)

                if len(df_all) >= header_index + 1:
                    top_row = df_all.iloc[0].replace({np.nan: None, 'NaN': None})
                    base_row = df_all.iloc[header_index].replace({np.nan: None, 'NaN': None})

                    combined_headers = []
                    current_suffix = None
                    for i in range(len(top_row)):
                        if top_row[i] is not None:
                            current_suffix = str(top_row[i])
                        base_name = str(base_row[i]) if base_row[i] is not None else f"col_{i}"
                        combined_headers.append(f"{base_name}_{current_suffix}" if current_suffix else base_name)

                    df = df_all.iloc[header_index + 1:]
                    df.columns = combined_headers
                    print(f"Processed trend table '{sheet_name}' with dynamic combined headers")
                else:
                    print(f"Sheet '{sheet_name}' lacks sufficient rows for combined header processing")
                    continue

            else:
                print(f"Skipping unrecognized sheet '{sheet_name}'")
                continue

            # Clean & Save
            df = df.reset_index(drop=True)
            df.columns = [str(col).replace(' ', '_').replace('(', '').replace(')', '').replace('.', '_')
                          .replace('-', '_').replace('/', '_').replace('\\', '_') for col in df.columns]
            table_name = ''.join(e for e in sheet_name if e.isalnum() or e == '_') or f"table_{sheet_name_lower.replace(' ', '_')}"
            df.to_sql(table_name, engine, index=False, if_exists='replace')
            print(f"✅ Imported '{sheet_name}' as '{table_name}' with {len(df)} rows.\n")

        except Exception as e:
            print(f"❌ Could not import sheet '{sheet_name}': {e}\n")

    return engine

def query_db(db_path, query_string):
    """
    Connects to the SQLite database and executes a given SQL query.
    Returns the query result.
    """
    engine = create_engine(f"sqlite:///{db_path}")
    with engine.connect() as connection:
        try:
            result = connection.execute(text(query_string))
            rows = result.fetchall()
            # Get column names for better display if needed
            # column_names = result.keys()
            # return [dict(zip(column_names, row)) for row in rows] # Returns list of dicts
            return rows # Returns list of tuples
        except Exception as e:
            return f"Error executing query: {e}"

def get_db_schema(db_path):
    """
    Returns the schema of the database (table names and their columns).
    """
    engine = create_engine(f"sqlite:///{db_path}")
    schema = {}
    with engine.connect() as connection:
        # Get table names
        table_names_query = text("SELECT name FROM sqlite_master WHERE type='table';")
        tables_result = connection.execute(table_names_query)
        table_names = [row[0] for row in tables_result.fetchall()]

        for table_name in table_names:
            # Get column info for each table
            columns_query = text(f"PRAGMA table_info({table_name});")
            columns_result = connection.execute(columns_query)
            columns = []
            for col_info in columns_result.fetchall():
                # col_info format: (cid, name, type, notnull, dflt_value, pk)
                columns.append(f"{col_info[1]} ({col_info[2]})")
            schema[table_name] = columns
    return schema
